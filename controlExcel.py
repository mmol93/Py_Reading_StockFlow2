from openpyxl import load_workbook
import getpass
import datetime

## 불러올 엑셀 파일 설정

# 현재 컴퓨터의 유저명 들고오기
userName = getpass.getuser()

# oneDrive 주소 가져오기
path1 = "C:/Users/"
path2 = "/OneDrive/AI_List.xlsx"

path = path1 + str(userName) + path2

# data_only=True : 수식없이 값만 가져오게 설정
# 내가 저장한 AI_List 들고오기
load_wb = load_workbook(path, data_only=True)

# 불러올 시트 설정
load_sheet = load_wb['평균선']

StockList = []  # 인공지능 종목 리스트
webPageList = []  # 인공지능 추천가

## *** 엑셀 데이터 저장
def saveExcell(load_wb):
    # 엑셀이 열려있을 때 저장하려고 할 경우 에러 메시지 출력
    try:
        excell_file = load_wb
        excell_file.save(path)
    except PermissionError:
        print("열려있는 AI_List.xlsx 엑셀을 닫으세요")

# 셀에 종목 기록하기(코스피, 코스닥만)
# 매개변수 설명
# 1. 코스닥인지 코스피인지 명시
# 2~5. 5일, 20일, 60일, 120일 데이터 넣기
def writeIndex(index,
               average5,
               average20,
               average60,
               average120):
    today = datetime.datetime.now()
    today = today.strftime("%Y-%m-%d")
    load_sheet.cell(3, 2).value = str(today)

    if index == "kospi":
        # 'C5'~'F5'까지 적어넣기
        load_sheet.cell(5, 3).value = str(average5)
        load_sheet.cell(5, 4).value = str(average20)
        load_sheet.cell(5, 5).value = str(average60)
        load_sheet.cell(5, 6).value = str(average120)

    elif index == "kosdaq":
        # 'C6'~'F6'까지 적어넣기
        load_sheet.cell(6, 3).value = str(average5)
        load_sheet.cell(6, 4).value = str(average20)
        load_sheet.cell(6, 5).value = str(average60)
        load_sheet.cell(6, 6).value = str(average120)

    # 엑셀 저장
    saveExcell(load_wb)
